from flask import Flask, request, render_template
import os
import json
import re
from email import policy
from email.parser import BytesParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook

app = Flask(__name__)

# Set upload folder and allowed extensions
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'txt', 'eml'}

# Ensure the upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Function to check allowed file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Route for uploading files
@app.route('/')
def upload_file():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload():
    # Check if the post request has the file part
    if 'file' not in request.files:
        return render_template('upload.html', error='No file part')
    
    file = request.files['file']

    # If the user does not select a file, the browser may submit an empty part without filename
    if file.filename == '':
        return render_template('upload.html', error='No selected file')
    
    if file and allowed_file(file.filename):
        # Save the uploaded file
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        # Check if the file is .eml or .txt and call corresponding method
        if file.filename.endswith('.eml'):
            parsed_data = parse_eml_file(filepath)
        elif file.filename.endswith('.txt'):
            parsed_data = parse_txt_file(filepath)
        else:
            return render_template('upload.html', error='Unsupported file format')

        # Save parsed data to Excel
        excel_filename = save_to_excel(parsed_data)

        # Display success message
        success_message = f"File processed successfully! Added to: {excel_filename}"
        return render_template('upload.html', success=success_message)

    return render_template('upload.html', error='File type not allowed')

# Parsing function for .eml files
def parse_eml_file(filepath):
    parsed_data = {}

    # Parse the .eml file using BytesParser
    with open(filepath, 'rb') as f:
        msg = BytesParser(policy=policy.default).parse(f)

    # Extract headers
    parsed_data['Date'] = msg['Date']
    parsed_data['Subject'] = msg['Subject']
    parsed_data['From'] = msg['From']
    parsed_data['X-From'] = msg['X-From']
    parsed_data['To'] = msg['To']
    parsed_data['Reply-To'] = msg['Reply-To']
    parsed_data['Message-ID'] = msg['Message-ID']

    # Extract the email body
    parsed_data['Unstructured-Text'] = extract_email_body(msg)

    # Extract domain (mail server) from the "From" field
    if 'From' in parsed_data and parsed_data['From']:
        email_address = parsed_data['From']
        domain = email_address.split('@')[-1] if '@' in email_address else ''
        organization = domain.split('.')[0] if domain else ''
        parsed_data['Organization'] = organization

    return parsed_data

def extract_email_body(msg):
    if msg.is_multipart():
        # Search for the plain text part first
        for part in msg.iter_parts():
            if part.get_content_type() == 'text/plain':
                return part.get_payload(decode=True).decode(part.get_content_charset(), 'ignore').strip()

    # If it's not multipart or no plain text part is found, return the fallback payload
    return msg.get_payload(decode=True).decode(msg.get_content_charset(), 'ignore').strip() if msg else ""

# Parsing function for .txt files with JSON-like structure
def parse_txt_file(filepath):
    parsed_data = {}

    # Read the file and parse it as JSON
    with open(filepath, 'r') as f:
        json_data = json.load(f)
    
    # Extract relevant fields
    message = json_data.get('message', '')
    
    parsed_data['Message-ID'] = extract_field(message, 'Message-ID')
    parsed_data['Date'] = extract_field(message, 'Date')
    parsed_data['From'] = extract_field(message, 'From')
    parsed_data['X-From'] = extract_field(message, 'X-From')
    parsed_data['Subject'] = extract_field(message, 'Subject')
    parsed_data['Unstructured-Text'] = extract_body_from_message(message)

    # Extract domain (mail server) from the "From" field
    if 'From' in parsed_data and parsed_data['From']:
        email_address = parsed_data['From']
        domain = email_address.split('@')[-1] if '@' in email_address else ''
        organization = domain.split('.')[0] if domain else ''
        parsed_data['Organization'] = organization
    return parsed_data

# Helper function to extract fields using regex
def extract_field(message, field_name):
    match = re.search(rf'{field_name}: (.+)', message)
    return match.group(1).strip() if match else ''

# Extract body from the message (after the headers)
def extract_body_from_message(message):
    body_start_index = message.find('\n\n')  # Look for the first occurrence of double newline
    return message[body_start_index + 2:].strip() if body_start_index != -1 else ""

# Function to save parsed data to an Excel file
def save_to_excel(parsed_data):
    directory = 'FlaskApp_excel'
    
    # Ensure the directory exists
    os.makedirs(directory, exist_ok=True)
    
    excel_filename = os.path.join('parsed_email_details.xlsx')
    
    # Check if the file exists
    if os.path.exists(excel_filename):
        # Load the existing workbook
        wb = load_workbook(excel_filename)
        ws = wb.active
    else:
        # Create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        headers = ["Message-ID", "Date", "From", "Name", "Organization", "Subject", "Body"]
        ws.append(headers)

    # Append the new data
    ws.append([
        parsed_data.get('Message-ID', ''),
        parsed_data.get('Date', ''),
        parsed_data.get('From', ''),
        parsed_data.get('X-From', ''),
        parsed_data.get('Organization', ''),
        parsed_data.get('Subject', ''),
        parsed_data.get('Unstructured-Text', '')
    ])

    # Wrap text in all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Auto-adjust column widths based on the longest line in each cell
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 100)  # Limit max width to 100
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook
    wb.save(excel_filename)
    return excel_filename  # Return the filename for success message

if __name__ == '__main__':
    app.run(debug=True)