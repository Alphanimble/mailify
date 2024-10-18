import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from pymongo import MongoClient
import re

# Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['emails']
collection = db['complete_details']

# Specify the number of documents to retrieve
n = 10 # Change this number to retrieve different amounts of documents

# Retrieve the documents
documents = collection.find().limit(n)  # Adjust the number of documents to retrieve

# Create Excel workbook and worksheet
wb = Workbook()
ws = wb.active

# Define headers
headers = ["Message-ID", "Date", "From", "Name", "Subject", "Body"]

# Append headers to the worksheet
ws.append(headers)

# Set the font and alignment for the header row
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

for cell in ws[1]:  # ws[1] refers to the first row (header row)
    cell.font = header_font
    cell.alignment = header_alignment

# Function to safely extract fields using regex
def safe_extract(pattern, content):
    match = re.search(pattern, content)
    return match.group(1) if match else "N/A"  # Return "N/A" if no match found

# Iterate through the documents
for doc in documents:
    # Convert the document to a JSON string
    document_detail = json.dumps(doc, default=str)
    
    # Load JSON
    data = json.loads(document_detail)
    
    # Extract message content
    message_content = data['message']

    # Use the safe_extract function to extract desired fields
    message_id = safe_extract(r"Message-ID: (.+)", message_content)
    date = safe_extract(r"Date: (.+)", message_content)
    from_address = safe_extract(r"From: (.+)", message_content)
    subject = safe_extract(r"Subject: (.+)", message_content)
    Name = safe_extract(r"X-From: (.+)", message_content)
    body = re.split(r'\n\n', message_content, 1)[1].strip() if "\n\n" in message_content else "N/A"

    # Append document details as a new row in the worksheet
    ws.append([message_id, date, from_address, Name, subject, body])

# Wrap text in all cells and auto adjust column width
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center')

def get_column_width(text):
    lines = str(text).split('\n')
    return max(len(line) for line in lines)

# Auto-adjust column widths based on the longest line in each cell
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, get_column_width(cell.value))

    adjusted_width = min(max_length + 2, 100)  # Limit max width to 100
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the Excel workbook
wb.save("Multiple_Documents_Details2.xlsx")
