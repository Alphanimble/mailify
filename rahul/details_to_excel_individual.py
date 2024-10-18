import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pymongo import MongoClient
import re

def get_column_width(text):
    lines = str(text).split('\n')
    return max(len(line) for line in lines)

# Connect to MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['emails']
collection = db['complete_details']

# Specify which document to retrieve
n = 9  # Change this number to retrieve different documents

# Retrieve the document
document = collection.find().skip(n - 1).limit(1)

# Iterate through the documents
for doc in document:
    # Convert the document to a JSON string
    document_detail = json.dumps(doc, default=str)
    # Load JSON
    data = json.loads(document_detail)

message_content = data['message']

message_id = re.search(r"Message-ID: (.+)", message_content).group(1)
date = re.search(r"Date: (.+)", message_content).group(1)
from_address = re.search(r"From: (.+)", message_content).group(1)
subject = re.search(r"Subject: (.+)", message_content).group(1)
Name = re.search(r"X-From: (.+)", message_content).group(1)
body = re.split(r'\n\n', message_content, 1)[1].strip()

wb = Workbook()
ws = wb.active

headers = ["Message-ID", "Date", "From", "Name", "Subject", "Body"]

ws.append(headers)

ws.append([message_id, date, from_address, Name, subject, body])

# Wrap text in all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center')

# Auto-adjust column widths based on the longest line in each cell
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, get_column_width(cell.value))

    adjusted_width = min(max_length + 2, 100)  # Limit max width to 100
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save("Individual_Details.xlsx")